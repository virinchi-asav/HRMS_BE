"""Ports UserController.php.

Deliberate bug fixes (per the "fix obvious bugs" decision):
- updateProfile (self-service) called dd($validator->errors()) on validation failure -
  a debug halt. FastAPI's own request validation (422) replaces this naturally since we
  use typed Pydantic schemas instead of manual Validator::make() calls.
- The self-service profile update form accepted `role` in its validated field set,
  meaning any logged-in user could set their own role to Admin. That field is not
  settable via update_own_profile here at all - only the admin-facing
  update_user_profile_as_admin path (already role-gated to Admin/HR in the router) can
  change a user's role/can_update/status.
"""

import logging
from datetime import datetime, date

from fastapi import UploadFile
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.hrms.core.config import hrms_settings
from app.hrms.core.constants import DEFAULT_PASSWORD, Role
from app.hrms.core.security import hash_password, verify_password
from app.hrms.models.certificate import CertificateTemplateEntity, TrainingCertificateEntity
from app.hrms.models.task_assessment import BankQuestionEntity, QuestionBankEntity, TaskAssigneeEntity, TaskEntity
from app.hrms.models.training import (
    TrainingAssessmentEntity,
    TrainingCommentEntity,
    TrainingDayEntryEntity,
    TrainingMaterialEntity,
    TrainingProgramEntity,
    TrainingTraineeEntity,
)
from app.hrms.models.user import UserEntity
from app.hrms.schemas.user import UserAdminUpdateRequest, UserCreateRequest, UserProfileUpdateRequest
from app.hrms.services import email_service, file_storage_service
from app.hrms.services.email_service import send_email
from app.services.account_service import get_or_create_bench_account
from app.utils.pagination import PageResult, paginate

# Every FK relationship to users.id that would raise a raw IntegrityError on hard
# delete (verified against information_schema.KEY_COLUMN_USAGE) - checked up front so
# delete_user can raise one clear, itemized message instead of a 500.
_DELETE_BLOCKER_CHECKS = [
    (TrainingProgramEntity, TrainingProgramEntity.trainer_id, "the Trainer on {n} training program(s)"),
    (TrainingProgramEntity, TrainingProgramEntity.bu_head_id, "the BU Head on {n} training program(s)"),
    (TrainingProgramEntity, TrainingProgramEntity.created_by, "the creator of {n} training program(s)"),
    (TrainingTraineeEntity, TrainingTraineeEntity.trainee_id, "a Trainee on {n} training program(s)"),
    (TrainingDayEntryEntity, TrainingDayEntryEntity.created_by, "the logger of {n} training day-entry(ies)"),
    (TrainingCommentEntity, TrainingCommentEntity.author_id, "the author of {n} training comment(s)"),
    (TrainingMaterialEntity, TrainingMaterialEntity.added_by, "the uploader of {n} training material(s)"),
    (TrainingAssessmentEntity, TrainingAssessmentEntity.trainee_id, "the Trainee on {n} training assessment(s)"),
    (TrainingAssessmentEntity, TrainingAssessmentEntity.created_by, "the creator of {n} training assessment(s)"),
    (TrainingAssessmentEntity, TrainingAssessmentEntity.reviewed_by, "the reviewer of {n} training assessment(s)"),
    (TrainingCertificateEntity, TrainingCertificateEntity.trainee_id, "the recipient of {n} certificate(s)"),
    (TrainingCertificateEntity, TrainingCertificateEntity.issued_by, "the issuer of {n} certificate(s)"),
    (CertificateTemplateEntity, CertificateTemplateEntity.uploaded_by, "the uploader of {n} certificate template(s)"),
    (QuestionBankEntity, QuestionBankEntity.created_by, "the creator of {n} question bank(s)"),
    (BankQuestionEntity, BankQuestionEntity.created_by, "the author of {n} bank question(s)"),
    (TaskEntity, TaskEntity.created_by, "the creator of {n} Task Assessment(s)"),
    (TaskAssigneeEntity, TaskAssigneeEntity.trainee_id, "assigned to {n} Task Assessment(s)"),
]


async def _delete_blockers(db: AsyncSession, user_id: int) -> list[str]:
    blockers = []
    for model, column, phrase in _DELETE_BLOCKER_CHECKS:
        count = (await db.execute(select(func.count()).select_from(model).where(column == user_id))).scalar_one()
        if count:
            blockers.append(phrase.format(n=count))
    return blockers

logger = logging.getLogger(__name__)

FILE_FIELDS = [
    "passport_photo",
    "latest_cv",
    "aadhar_card",
    "pan_card",
    "marksheet_10th",
    "marksheet_12th",
    "sem_1",
    "sem_2",
    "sem_3",
    "sem_4",
    "sem_5",
    "sem_6",
    "sem_7",
    "sem_8",
    "pg_sem_1",
    "pg_sem_2",
    "pg_sem_3",
    "pg_sem_4",
    "consolidated_marksheet",
    "pg_consolidated_marksheet",
]

# Fields update_user_profile_as_admin (Admin/HR editing someone's profile) may set that
# update_own_profile (self-service) must never touch.
ADMIN_ONLY_PROFILE_FIELDS = {
    "role",
    "status",
    "can_update",
    "admin_message",
    "reviewed_by",
    "reviewed_at",
    "kms_department_id",
    "kms_account_id",
    "kms_user_type_id",
}


async def get_user(db: AsyncSession, user_id: int) -> UserEntity | None:
    return await db.get(UserEntity, user_id)


async def get_by_email(db: AsyncSession, email: str) -> UserEntity | None:
    result = await db.execute(select(UserEntity).where(UserEntity.email == email))
    return result.scalar_one_or_none()


async def authenticate(db: AsyncSession, email: str, password: str) -> UserEntity | None:
    user = await get_by_email(db, email)
    if user is None or user.deleted_at is not None:
        return None
    if not verify_password(password, user.password):
        return None
    return user


async def admin_list_users(
    db: AsyncSession,
    page_number: int,
    page_size: int,
    role: int | None,
    search: str | None,
    account_id: int | None = None,
    work_location: str | None = None,
) -> PageResult:
    """Mirrors UserController::index - role IN (1,2,4,5,6,7), i.e. Employee(3) is
    intentionally excluded from this admin list; HR(2) is included so HR accounts are
    visible/manageable in the Users Grid.

    account_id filters to users whose own KMS Account Type (kms_account_id) matches -
    used e.g. by the Task Assessment trainee picker to only show trainees under a given
    account once an Account Type has been chosen.

    work_location filters to users at a given office location (see hrms_settings.work_locations) -
    used by the Task Assessment "assign to all employees at this location" feature."""
    stmt = select(UserEntity).where(
        UserEntity.role.in_([Role.ADMIN, Role.HR, Role.CANDIDATE, Role.MANAGER, Role.TEAM_MEMBER, Role.BU_HEAD])
    )
    if search:
        stmt = stmt.where(or_(UserEntity.name.ilike(f"%{search}%"), UserEntity.email.ilike(f"%{search}%")))
    if role:
        stmt = stmt.where(UserEntity.role == role)
    if account_id:
        stmt = stmt.where(UserEntity.kms_account_id == account_id)
    if work_location:
        stmt = stmt.where(UserEntity.work_location == work_location)
    stmt = stmt.order_by(UserEntity.id.desc())
    return await paginate(db, stmt, page_number, page_size)


async def list_team_member_ids_by_location(db: AsyncSession, work_location: str) -> list[int]:
    """All active Team Member user ids at a given work location - unpaginated, used to
    bulk-assign a Task Assessment to every employee at that location."""
    result = await db.execute(
        select(UserEntity.id).where(UserEntity.role == Role.TEAM_MEMBER, UserEntity.work_location == work_location)
    )
    return list(result.scalars().all())


async def admin_create_user(db: AsyncSession, kms_db: AsyncSession, data: UserCreateRequest) -> UserEntity:
    entity = UserEntity(
        name=data.name,
        email=data.email,
        password=hash_password(hrms_settings.hrms_default_password),
        role=data.role,
        can_update=True,
    )
    # The quick-add form never collects an account/department - default a freshly
    # registered user to the "Bench" account type until HR later allocates them to a
    # real client account via the profile editor.
    if entity.kms_account_id is None and entity.kms_department_id is None:
        bench_account = await get_or_create_bench_account(kms_db)
        entity.kms_account_id = bench_account.account_id

    db.add(entity)
    await db.commit()
    await db.refresh(entity)

    await email_service.send_onboarding_email(data.email, data.name, hrms_settings.hrms_default_password)
    return entity


async def admin_update_user(db: AsyncSession, user_id: int, data: UserAdminUpdateRequest) -> UserEntity | None:
    entity = await get_user(db, user_id)
    if entity is None:
        return None
    if data.name is not None:
        entity.name = data.name
    if data.email is not None:
        entity.email = data.email
    if data.role is not None:
        entity.role = data.role
    if data.password:
        entity.password = hash_password(data.password)
    await db.commit()
    await db.refresh(entity)
    return entity


async def delete_user(db: AsyncSession, user_id: int) -> bool:
    """Mirrors UserController::destroy: delete the user's skills, then hard-delete the
    user itself (bypassing the `deleted_at` column that exists on the table but is
    never actually used for user rows in the source app).

    Raises ValueError (-> 400) with an itemized list if the user is still referenced
    elsewhere (trainer/BU-head on a training, task/bank creator, etc.) - without this
    check the hard delete below would instead fail with a raw, unhandled FK
    IntegrityError (a 500) that doesn't tell the caller what to fix."""
    from app.hrms.models.skill import SkillEntity

    entity = await get_user(db, user_id)
    if entity is None:
        return False

    blockers = await _delete_blockers(db, user_id)
    if blockers:
        raise ValueError(
            f"Cannot delete this user - they are still referenced as: {'; '.join(blockers)}. "
            "Remove or reassign those records first."
        )

    skills = (await db.execute(select(SkillEntity).where(SkillEntity.user_id == user_id))).scalars().all()
    for skill in skills:
        await db.delete(skill)

    await db.delete(entity)
    await db.commit()
    return True


async def _apply_uploaded_files(user_id: int, files: dict[str, UploadFile | None]) -> dict[str, str]:
    """Mirrors the `$fileFields` loop in updateProfile/updateUserProfile: saves each
    provided file as "{field}_{timestamp}.{ext}" under uploads/users/{id}/{field}/."""
    stored: dict[str, str] = {}
    for field, file in files.items():
        if file is None:
            continue
        ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin"
        stored_name = f"{field}_{int(datetime.utcnow().timestamp())}.{ext}"
        content = await file.read()
        await file_storage_service.save_file(file_storage_service.user_document_dir(user_id, field), stored_name, content)
        stored[field] = stored_name
    return stored


async def update_own_profile(
    db: AsyncSession, current_user: UserEntity, data: UserProfileUpdateRequest, files: dict[str, UploadFile | None]
) -> UserEntity:
    """Self-service profile edit (UserController::updateProfile). Always operates on
    the authenticated user - `data.role` (and any other admin-only field) is ignored
    even if the client sends it; see ADMIN_ONLY_PROFILE_FIELDS."""
    updates = data.model_dump(exclude_unset=True, exclude=ADMIN_ONLY_PROFILE_FIELDS)
    password = updates.pop("password", None)

    for field, value in updates.items():
        if field == "email" and value is None:
            continue
        setattr(current_user, field, value)

    if password:
        current_user.password = hash_password(password)

    stored_files = await _apply_uploaded_files(current_user.id, files)
    for field, filename in stored_files.items():
        setattr(current_user, field, filename)

    current_user.modified_by = current_user.id
    current_user.modified_time = datetime.utcnow()
    current_user.can_update = False
    current_user.status = "pending"

    await db.commit()
    await db.refresh(current_user)

    notify_email = hrms_settings.hrms_profile_update_notify_email
    if notify_email:
        html = f"<p>{current_user.name} updated their profile and it is pending review.</p>"
        await send_email(notify_email, "User Profile Updated", html)

    return current_user


async def update_user_profile_as_admin(
    db: AsyncSession,
    target_user_id: int,
    data: UserProfileUpdateRequest,
    files: dict[str, UploadFile | None],
) -> UserEntity | None:
    """Admin/HR editing a specific user's full profile (UserController::
    updateUserProfile) - the only path allowed to change role/training/project-
    allocation fields."""
    entity = await get_user(db, target_user_id)
    if entity is None:
        return None

    updates = data.model_dump(exclude_unset=True)
    password = updates.pop("password", None)

    previous_onboarding_date = entity.mks_onboarding_date
    new_onboarding_date = updates.get("mks_onboarding_date", previous_onboarding_date)

    for field, value in updates.items():
        setattr(entity, field, value)

    if password:
        entity.password = hash_password(password)

    stored_files = await _apply_uploaded_files(entity.id, files)
    for field, filename in stored_files.items():
        setattr(entity, field, filename)

    entity.modified_by = target_user_id
    entity.modified_time = datetime.utcnow()
    entity.can_update = False

    if previous_onboarding_date != new_onboarding_date:
        manager = await get_user(db, int(entity.bu_head)) if entity.bu_head and entity.bu_head.isdigit() else None
        cc = manager.email if manager and manager.email else None
        html = (
            f"<p>{entity.name}'s onboarding date has been set to {new_onboarding_date}.</p>"
        )
        await send_email(hrms_settings.hrms_onboarding_notify_email, "Onboarding Update", html, cc=cc)

    await db.commit()
    await db.refresh(entity)
    return entity


async def get_reporting_candidates(db: AsyncSession) -> list[UserEntity]:
    """Mirrors showProfile's $reportingUsers dropdown - role in [Admin, Manager, BU Head]."""
    result = await db.execute(select(UserEntity).where(UserEntity.role.in_([Role.ADMIN, Role.MANAGER, Role.BU_HEAD])))
    return list(result.scalars().all())


STATIC_ACCOUNT_OPTIONS = ["Outside MKS", "Theoretical/Online-Course"]


async def get_accounts(db: AsyncSession) -> list[str]:
    """Distinct project_name values from users, plus the fixed options that were always
    present in the legacy Laravel form - shared by the Skills "Account" dropdown and the
    User Profile "Project Allocation > Account" dropdown so both stay in sync."""
    result = await db.execute(select(UserEntity.project_name).where(UserEntity.project_name.is_not(None)).distinct())
    return sorted(set(result.scalars().all()) | set(STATIC_ACCOUNT_OPTIONS))


def get_work_locations() -> list[str]:
    """The fixed set of office locations (see HRMS_WORK_LOCATIONS in .env) offered on
    the Work Location dropdown and the Task Assessment "assign by location" feature."""
    return hrms_settings.work_locations


ROLE_NAME_TO_ID = {
    "admin": Role.ADMIN,
    "team lead": Role.MANAGER,
    "manager": Role.MANAGER,
    "bu head": Role.BU_HEAD,
}


async def _save_user_data(db: AsyncSession, data: dict, user: UserEntity | None) -> UserEntity:
    """Mirrors UserController::saveUserData exactly, including the hard-coded BU Head
    employee-id overrides (see app/hrms/utils/reporting.py)."""
    from app.hrms.utils.reporting import FORCED_BU_HEAD_EMPLOYEE_IDS, extract_employee_id

    if user is None:
        user = UserEntity(email=data["email"], password=hash_password(data.get("password") or DEFAULT_PASSWORD))
        db.add(user)

    if data.get("name"):
        user.name = data["name"]
    user.email = data["email"]

    role_name = str(data.get("role", "")).strip().lower()
    user.role = ROLE_NAME_TO_ID.get(role_name, Role.TEAM_MEMBER)

    for field in (
        "employee_id",
        "gender",
        "marital_status",
        "birth_date",
        "wedding_day",
        "father_name",
        "emergency_contact_person_name",
        "emergency_contact_number",
        "work_location",
        "project_name",
        "source_of_hire",
        "job_title",
        "department",
        "total_experience",
        "experience",
        "date_of_joining",
        "employee_type",
        "employee_status",
        "job_description",
        "ask_me_about",
        "about_me",
    ):
        if data.get(field) is not None:
            setattr(user, field, data[field])

    raw_reporting_to = data.get("reporting_to")
    if raw_reporting_to:
        employee_id = extract_employee_id(raw_reporting_to)
        if employee_id:
            result = await db.execute(select(UserEntity).where(UserEntity.employee_id == employee_id))
            reporting_user = result.scalar_one_or_none()
            if reporting_user:
                if reporting_user.role not in (Role.MANAGER, Role.BU_HEAD):
                    reporting_user.role = (
                        Role.BU_HEAD if reporting_user.employee_id in FORCED_BU_HEAD_EMPLOYEE_IDS else Role.MANAGER
                    )
                user.reporting_to = str(reporting_user.id)

    if data.get("password"):
        user.password = hash_password(data["password"])

    await db.flush()
    return user


async def sync_user_from_webhook(db: AsyncSession, data: dict) -> tuple[UserEntity, bool]:
    """Mirrors UserController::webhookUserSync."""
    existing = await get_by_email(db, data["email"])
    is_new = existing is None

    if is_new and not data.get("name"):
        raise ValueError("The name field is required when creating a new user.")

    onboarding_password = None
    if is_new:
        onboarding_password = data.get("password") or DEFAULT_PASSWORD
        data = {**data, "password": onboarding_password}

    user = await _save_user_data(db, data, existing)
    await db.commit()
    await db.refresh(user)

    if is_new:
        await email_service.send_onboarding_email(user.email, user.name, onboarding_password)

    return user, is_new


EXPORT_COLUMNS = [
    "name", "email", "role", "employee_id", "first_name", "middle_name", "last_name",
    "added_by", "added_time", "onboarding_status", "mobile_phone", "personal_email",
    "modified_by", "modified_time", "work_location", "project_name", "skillset",
    "reporting_to", "source_of_hire", "seating_location", "job_role", "total_experience",
    "experience", "band", "department", "job_title", "date_of_joining",
    "employee_experience", "probation_end_date", "probation_status", "employee_type",
    "employee_status", "work_phone", "extension", "final_rating", "present_address",
    "father_name", "birth_date", "gender", "marital_status", "wedding_day", "citizenship",
    "permanent_address", "blood_group", "age", "pan_card_number", "aadhaar_card_number",
    "passport_number", "uan_number", "emergency_contact_person_name",
    "emergency_contact_number", "candidate_type", "emergency_contact_person_relation",
    "job_description", "ask_me_about", "about_me", "tenth_school_name", "tenth_board",
    "tenth_field_of_study", "tenth_date_of_completion", "twelfth_school_name",
    "twelfth_board", "twelfth_field_of_study", "twelfth_date_of_completion",
    "graduation_school_name", "graduation_degree", "graduation_field_of_study",
    "graduation_date_of_completion", "post_graduation_school_name",
    "post_graduation_degree", "post_graduation_field_of_study",
    "post_graduation_date_of_completion",
]

EXPORT_HEADINGS = [
    "Name", "Email", "Role", "Employee ID", "First Name", "Middle Name", "Last Name",
    "Added By", "Added Time", "Onboarding Status", "Mobile Phone", "Personal Email",
    "Modified By", "Modified Time", "Work Location", "Project Name", "Skillset",
    "Reporting To", "Source of Hire", "Seating Location", "Job Role", "Total Experience",
    "Relevant Experience", "Band", "Department", "Job Title", "Date of Joining",
    "Employee Experience", "Probation End Date", "Probation Status", "Employee Type",
    "Employee Status", "Work Phone", "Extension", "Final Rating", "Present Address",
    "Father's Name", "Date of Birth", "Gender", "Marital Status", "Wedding Day",
    "Citizenship", "Permanent Address", "Blood Group", "Age", "PAN Card Number",
    "Aadhaar Card Number", "Passport Number", "UAN Number",
    "Emergency Contact Person Name", "Emergency Contact Number", "Candidate Type",
    "Emergency Contact Person Relation", "Job Description", "Ask Me About", "About Me",
    "10th School Name", "10th Board", "10th Field of Study", "10th Date of Completion",
    "12th School Name", "12th Board", "12th Field of Study", "12th Date of Completion",
    "Graduation School Name", "Graduation Degree", "Graduation Field of Study",
    "Graduation Date of Completion", "Post Graduation School Name",
    "Post Graduation Degree", "Post Graduation Field of Study",
    "Post Graduation Date of Completion",
]

ROLE_ID_TO_NAME = {1: "Admin", 2: "HR", 3: "Employee", 4: "Candidate"}


async def export_candidates_xlsx(db: AsyncSession) -> bytes:
    """Mirrors UsersExport - candidates (role=4) only, full column set."""
    import io

    import openpyxl

    result = await db.execute(select(UserEntity).where(UserEntity.role == Role.CANDIDATE))
    users = result.scalars().all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(EXPORT_HEADINGS)
    for user in users:
        row = []
        for col in EXPORT_COLUMNS:
            value = ROLE_ID_TO_NAME.get(user.role, "Unknown") if col == "role" else getattr(user, col, None)
            row.append(str(value) if value is not None else None)
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def import_users_xlsx(db: AsyncSession, content: bytes) -> dict:
    """Mirrors UsersImport: only name+email read from the sheet; skip if an identical
    user already exists, error if the email exists with a different name, otherwise
    create with role=Candidate and the default password."""
    import io

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.worksheets[0]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_idx = headers.index("name") if "name" in headers else None
    email_idx = headers.index("email") if "email" in headers else None

    created, skipped, errors = 0, 0, []
    if name_idx is None or email_idx is None:
        return {"created": 0, "skipped": 0, "errors": ["Sheet must have 'name' and 'email' columns."]}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[name_idx] if name_idx < len(row) else None
        email = row[email_idx] if email_idx < len(row) else None
        if not name or not email:
            continue

        existing = await get_by_email(db, email)
        if existing and existing.name == name:
            skipped += 1
            continue
        if existing:
            errors.append(f"The email {email} already exists in the database.")
            continue

        user = UserEntity(name=name, email=email, password=hash_password(DEFAULT_PASSWORD), role=Role.CANDIDATE)
        db.add(user)
        await db.flush()
        await email_service.send_onboarding_email(email, name, DEFAULT_PASSWORD)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


async def delete_user_by_employee_id(db: AsyncSession, employee_id: str) -> bool:
    from app.hrms.models.skill import SkillEntity

    result = await db.execute(select(UserEntity).where(UserEntity.employee_id == employee_id))
    user = result.scalar_one_or_none()
    if user is None:
        return False

    skills = (await db.execute(select(SkillEntity).where(SkillEntity.user_id == user.id))).scalars().all()
    for skill in skills:
        await db.delete(skill)
    await db.delete(user)
    await db.commit()
    return True
