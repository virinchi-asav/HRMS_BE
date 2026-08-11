import random
from datetime import datetime, timedelta

from pydantic import ValidationError as PydanticValidationError
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.hrms.core.constants import Role
from app.hrms.core.deps import CurrentHrmsUser
from app.hrms.models.task_assessment import (
    AssigneeStatus,
    BankQuestionEntity,
    BankQuestionOptionEntity,
    DifficultyLevel,
    QuestionBankEntity,
    QuestionMode,
    QuestionType,
    TaskAnswerEntity,
    TaskAssigneeEntity,
    TaskEntity,
    TaskQuestionEntity,
    TaskQuestionOptionEntity,
    TaskStatus,
)
from app.hrms.models.training import AssessmentGivenBy, TrainingProgramEntity, TrainingStatus, TrainingTraineeEntity
from app.hrms.models.user import UserEntity as HrmsUserEntity
from app.hrms.services import email_service, skill_service, user_service
from app.models.account import AccountEntity
from app.hrms.schemas.task_assessment import (
    AnswerSaveRequest,
    AnswerSaveResponse,
    BankCreateRequest,
    BankQuestionCreateRequest,
    BankQuestionImportError,
    BankQuestionImportResponse,
    BankQuestionResponse,
    BankQuestionUpdateRequest,
    BankResponse,
    BankUpdateRequest,
    OptionInput,
    OptionResponse,
    RetryResponse,
    TaskAssigneeSummary,
    TaskCreateRequest,
    TaskManageDetailResponse,
    TaskOptionManageItem,
    TaskOptionTakeItem,
    TaskQuestionManageItem,
    TaskQuestionTakeItem,
    TaskReportResponse,
    TaskReportRow,
    TaskResultResponse,
    TaskTakeResponse,
)
from app.utils.pagination import PageResult, paginate

_FINALIZED_STATUSES = (AssigneeStatus.SUBMITTED.value, AssigneeStatus.AUTO_SUBMITTED.value)

# Maps a Task's Difficulty Level to the Skills module's 1-5 Rating scale when a passing
# attempt auto-logs the skill (see _grade_and_finalize) - Rating 1 ("Cannot perform the
# task") is deliberately never assigned this way, since passing an assessment at any
# difficulty implies some real ability.
DIFFICULTY_TO_RATING = {
    DifficultyLevel.BEGINNER.value: "2",
    DifficultyLevel.INTERMEDIATE.value: "3",
    DifficultyLevel.ADVANCED.value: "4",
    DifficultyLevel.EXPERT.value: "5",
}


async def _user_names(hrms_db: AsyncSession, user_ids: set[int]) -> dict[int, str]:
    if not user_ids:
        return {}
    result = await hrms_db.execute(select(HrmsUserEntity.id, HrmsUserEntity.name).where(HrmsUserEntity.id.in_(user_ids)))
    return {row.id: row.name for row in result.all()}


def _can_manage(current_user: CurrentHrmsUser, created_by: int) -> bool:
    return created_by == current_user.id or current_user.role in (Role.ADMIN, Role.HR)


async def _training_trainee_ids(hrms_db: AsyncSession, training_id: int) -> set[int]:
    result = await hrms_db.execute(
        select(TrainingTraineeEntity.trainee_id).where(TrainingTraineeEntity.training_id == training_id)
    )
    return {row[0] for row in result.all()}


async def _get_latest_assignee(hrms_db: AsyncSession, task_id: int, trainee_id: int) -> TaskAssigneeEntity | None:
    """The trainee's current attempt - the highest attempt_number row for (task,
    trainee). attempt_number only ever increases (see TaskAssigneeEntity), so this is
    always the one live/most-recent attempt."""
    result = await hrms_db.execute(
        select(TaskAssigneeEntity)
        .where(TaskAssigneeEntity.task_id == task_id, TaskAssigneeEntity.trainee_id == trainee_id)
        .order_by(TaskAssigneeEntity.attempt_number.desc())
        .limit(1)
    )
    return result.scalar_one_or_none()


async def _latest_assignees_for_task(hrms_db: AsyncSession, task_id: int) -> list[TaskAssigneeEntity]:
    """One row per trainee ever assigned to this task - their current/latest attempt,
    collapsing any retry/reassign history down to what's live now."""
    result = await hrms_db.execute(
        select(TaskAssigneeEntity)
        .where(TaskAssigneeEntity.task_id == task_id)
        .order_by(TaskAssigneeEntity.trainee_id, TaskAssigneeEntity.attempt_number)
    )
    latest_by_trainee: dict[int, TaskAssigneeEntity] = {}
    for a in result.scalars().all():
        latest_by_trainee[a.trainee_id] = a  # ascending attempt_number - later rows overwrite earlier ones
    return list(latest_by_trainee.values())


async def _latest_assignees_for_tasks(hrms_db: AsyncSession, task_ids: list[int]) -> dict[int, list[TaskAssigneeEntity]]:
    """Multi-task equivalent of _latest_assignees_for_task, for list views."""
    if not task_ids:
        return {}
    result = await hrms_db.execute(
        select(TaskAssigneeEntity)
        .where(TaskAssigneeEntity.task_id.in_(task_ids))
        .order_by(TaskAssigneeEntity.task_id, TaskAssigneeEntity.trainee_id, TaskAssigneeEntity.attempt_number)
    )
    latest_by_task_trainee: dict[tuple[int, int], TaskAssigneeEntity] = {}
    for a in result.scalars().all():
        latest_by_task_trainee[(a.task_id, a.trainee_id)] = a
    by_task: dict[int, list[TaskAssigneeEntity]] = {}
    for (task_id, _trainee_id), a in latest_by_task_trainee.items():
        by_task.setdefault(task_id, []).append(a)
    return by_task


def _assignee_attempt_flags(assignee: TaskAssigneeEntity) -> tuple[bool, bool]:
    """(can_retry, locked_out) - only meaningful once this attempt is finalized and
    failed; a pass, or a still-open attempt, is neither. can_retry is self-service
    (trainee calls retry_task); locked_out means every attempt this trainee currently
    has is used up and only HR's reassign_task can grant more."""
    failed = assignee.status in _FINALIZED_STATUSES and assignee.passed is False
    if not failed:
        return False, False
    return assignee.attempt_number < assignee.max_attempts, assignee.attempt_number >= assignee.max_attempts


def _to_result_response(task_id: int, assignee: TaskAssigneeEntity) -> TaskResultResponse:
    can_retry, locked_out = _assignee_attempt_flags(assignee)
    return TaskResultResponse(
        task_id=task_id,
        status=assignee.status,
        marks_obtained=assignee.marks_obtained,
        total_marks=assignee.total_marks,
        percentage=assignee.percentage,
        passed=assignee.passed,
        submitted_at=assignee.submitted_at,
        submit_reason=assignee.submit_reason,
        attempt_number=assignee.attempt_number,
        max_attempts=assignee.max_attempts,
        can_retry=can_retry,
        locked_out=locked_out,
    )


def _grade_single_answer(
    task_question: TaskQuestionEntity,
    options_by_id: dict[int, TaskQuestionOptionEntity],
    selected_option_id: int | None,
    answer_text: str | None,
) -> tuple[bool, int]:
    if task_question.question_type == QuestionType.MULTIPLE_CHOICE.value:
        option = options_by_id.get(selected_option_id) if selected_option_id is not None else None
        is_correct = option is not None and option.is_correct
    else:
        correct = (task_question.correct_answer_text or "").strip().lower()
        given = (answer_text or "").strip().lower()
        is_correct = bool(given) and given == correct
    return is_correct, (task_question.marks if is_correct else 0)


async def _grade_and_finalize(
    hrms_db: AsyncSession, assignee: TaskAssigneeEntity, task: TaskEntity, status: str, finalized_at: datetime
) -> None:
    """Idempotent - a no-op if the assignee is already SUBMITTED/AUTO_SUBMITTED. Sums
    already-graded TaskAnswerEntity rows (graded incrementally at each autosave) rather
    than re-grading everything here."""
    if assignee.status in _FINALIZED_STATUSES:
        return

    task_questions, _ = await _task_questions_with_options(hrms_db, task.id, assignee.id)
    total_marks = sum(q.marks for q in task_questions)

    ans_result = await hrms_db.execute(select(TaskAnswerEntity).where(TaskAnswerEntity.assignee_id == assignee.id))
    marks_obtained = sum(a.marks_awarded or 0 for a in ans_result.scalars().all())

    percentage = round((marks_obtained / total_marks) * 100, 2) if total_marks else 0.0
    assignee.status = status
    assignee.submitted_at = finalized_at
    assignee.marks_obtained = marks_obtained
    assignee.total_marks = total_marks
    assignee.percentage = percentage
    assignee.passed = percentage >= task.pass_percentage
    await hrms_db.commit()

    if assignee.passed and task.skill_name:
        rating = DIFFICULTY_TO_RATING.get(task.difficulty_level, "2")
        await skill_service.auto_log_skill_from_task_pass(
            hrms_db, assignee.trainee_id, task.skill_name, task.skill_category, rating
        )

    # HR (the task's creator) is notified either way - a fail is exactly the outcome
    # they need to see to decide whether to reassign (see retry_task/reassign_task).
    hr_user = await hrms_db.get(HrmsUserEntity, task.created_by)
    trainee = await hrms_db.get(HrmsUserEntity, assignee.trainee_id)
    if hr_user and trainee:
        await email_service.send_task_result_email(
            hr_user.email,
            trainee.name,
            task.title,
            assignee.passed,
            assignee.percentage,
            assignee.marks_obtained,
            assignee.total_marks,
            task.id,
        )


async def _finalize_if_expired(hrms_db: AsyncSession, assignee: TaskAssigneeEntity, task: TaskEntity) -> None:
    """The server-side timeout safety net - called at the top of every function that
    touches an in-progress assignee, and for every row a report reads, so a trainee who
    never returns after their deadline is never left stale."""
    if assignee.status != AssigneeStatus.IN_PROGRESS.value:
        return
    if assignee.deadline_at is None or datetime.utcnow() <= assignee.deadline_at:
        return
    await _grade_and_finalize(hrms_db, assignee, task, AssigneeStatus.AUTO_SUBMITTED.value, assignee.deadline_at)


# ---------------------------------------------------------------------------
# Question banks
# ---------------------------------------------------------------------------


async def _bank_question_counts(hrms_db: AsyncSession, bank_ids: list[int]) -> dict[int, int]:
    if not bank_ids:
        return {}
    result = await hrms_db.execute(
        select(BankQuestionEntity.bank_id, func.count())
        .where(BankQuestionEntity.bank_id.in_(bank_ids), BankQuestionEntity.is_active.is_(True))
        .group_by(BankQuestionEntity.bank_id)
    )
    return dict(result.all())


async def _account_names(kms_db: AsyncSession, account_ids: set[int]) -> dict[int, str]:
    if not account_ids:
        return {}
    result = await kms_db.execute(select(AccountEntity).where(AccountEntity.account_id.in_(account_ids)))
    return {a.account_id: a.account_name for a in result.scalars().all()}


async def _to_bank_response(hrms_db: AsyncSession, kms_db: AsyncSession, entity: QuestionBankEntity) -> BankResponse:
    names = await _user_names(hrms_db, {entity.created_by})
    counts = await _bank_question_counts(hrms_db, [entity.id])
    account_name = None
    if entity.account_id:
        accounts = await _account_names(kms_db, {entity.account_id})
        account_name = accounts.get(entity.account_id)
    return BankResponse(
        id=entity.id,
        name=entity.name,
        description=entity.description,
        account_id=entity.account_id,
        account_name=account_name,
        custom_account_type=entity.custom_account_type,
        is_active=entity.is_active,
        created_by=entity.created_by,
        created_by_name=names.get(entity.created_by, "Unknown"),
        question_count=counts.get(entity.id, 0),
        created_at=entity.created_at,
        updated_at=entity.updated_at,
    )


async def create_bank(
    hrms_db: AsyncSession, kms_db: AsyncSession, current_user: CurrentHrmsUser, data: BankCreateRequest
) -> BankResponse:
    now = datetime.utcnow()
    entity = QuestionBankEntity(
        name=data.name,
        description=data.description,
        account_id=data.account_id,
        custom_account_type=data.custom_account_type,
        is_active=True,
        created_by=current_user.id,
        created_at=now,
        updated_at=now,
    )
    hrms_db.add(entity)
    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_response(hrms_db, kms_db, entity)


async def list_banks(
    hrms_db: AsyncSession,
    kms_db: AsyncSession,
    current_user: CurrentHrmsUser,
    page_number: int,
    page_size: int,
    search: str | None,
    account_id: int | None = None,
) -> PageResult:
    """Shared org-wide library - readable by every role the router lets through, no
    ownership filter; only mutation is ownership-gated."""
    stmt = select(QuestionBankEntity).where(QuestionBankEntity.is_active.is_(True))
    if search:
        stmt = stmt.where(QuestionBankEntity.name.ilike(f"%{search}%"))
    if account_id is not None:
        stmt = stmt.where(QuestionBankEntity.account_id == account_id)
    stmt = stmt.order_by(QuestionBankEntity.id.desc())
    page_result = await paginate(hrms_db, stmt, page_number, page_size)

    banks = page_result.items
    names = await _user_names(hrms_db, {b.created_by for b in banks})
    counts = await _bank_question_counts(hrms_db, [b.id for b in banks])
    accounts = await _account_names(kms_db, {b.account_id for b in banks if b.account_id})
    page_result.items = [
        {
            "id": b.id,
            "name": b.name,
            "description": b.description,
            "account_id": b.account_id,
            "account_name": accounts.get(b.account_id) if b.account_id else None,
            "custom_account_type": b.custom_account_type,
            "is_active": b.is_active,
            "created_by": b.created_by,
            "created_by_name": names.get(b.created_by, "Unknown"),
            "question_count": counts.get(b.id, 0),
            "created_at": b.created_at,
            "updated_at": b.updated_at,
        }
        for b in banks
    ]
    return page_result


async def get_bank(
    hrms_db: AsyncSession, kms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int
) -> BankResponse | None:
    entity = await hrms_db.get(QuestionBankEntity, bank_id)
    if entity is None:
        return None
    return await _to_bank_response(hrms_db, kms_db, entity)


async def update_bank(
    hrms_db: AsyncSession, kms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, data: BankUpdateRequest
) -> BankResponse | None:
    entity = await hrms_db.get(QuestionBankEntity, bank_id)
    if entity is None or not _can_manage(current_user, entity.created_by):
        return None
    if data.name is not None:
        entity.name = data.name
    if data.description is not None:
        entity.description = data.description
    # Account type is always resent as a unit (both fields together) so switching
    # between an existing account and a custom "Other" label never leaves the old
    # value dangling on the side not being used.
    entity.account_id = data.account_id
    entity.custom_account_type = data.custom_account_type
    entity.updated_at = datetime.utcnow()
    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_response(hrms_db, kms_db, entity)


async def set_bank_active(
    hrms_db: AsyncSession, kms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, is_active: bool
) -> BankResponse | None:
    entity = await hrms_db.get(QuestionBankEntity, bank_id)
    if entity is None or not _can_manage(current_user, entity.created_by):
        return None
    entity.is_active = is_active
    entity.updated_at = datetime.utcnow()
    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_response(hrms_db, kms_db, entity)


# ---------------------------------------------------------------------------
# Bank questions
# ---------------------------------------------------------------------------


async def _to_bank_question_response(hrms_db: AsyncSession, entity: BankQuestionEntity) -> BankQuestionResponse:
    result = await hrms_db.execute(
        select(BankQuestionOptionEntity)
        .where(BankQuestionOptionEntity.question_id == entity.id)
        .order_by(BankQuestionOptionEntity.display_order, BankQuestionOptionEntity.id)
    )
    options = result.scalars().all()
    return BankQuestionResponse(
        id=entity.id,
        bank_id=entity.bank_id,
        module_name=entity.module_name,
        question_type=entity.question_type,
        question_text=entity.question_text,
        marks=entity.marks,
        correct_answer_text=entity.correct_answer_text,
        is_active=entity.is_active,
        options=[
            OptionResponse(id=o.id, option_text=o.option_text, is_correct=o.is_correct, display_order=o.display_order)
            for o in options
        ],
        created_at=entity.created_at,
        updated_at=entity.updated_at,
    )


async def add_question(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, data: BankQuestionCreateRequest
) -> BankQuestionResponse | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    if bank is None or not _can_manage(current_user, bank.created_by):
        return None
    now = datetime.utcnow()
    entity = BankQuestionEntity(
        bank_id=bank_id,
        module_name=data.module_name,
        question_type=data.question_type,
        question_text=data.question_text,
        marks=data.marks,
        correct_answer_text=data.correct_answer_text,
        is_active=True,
        created_by=current_user.id,
        created_at=now,
        updated_at=now,
    )
    hrms_db.add(entity)
    await hrms_db.flush()
    for idx, opt in enumerate(data.options):
        hrms_db.add(
            BankQuestionOptionEntity(question_id=entity.id, option_text=opt.option_text, is_correct=opt.is_correct, display_order=idx)
        )
    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_question_response(hrms_db, entity)


# ---------------------------------------------------------------------------
# Bulk question import (CSV/XLSX) + template
# ---------------------------------------------------------------------------

IMPORT_TEMPLATE_COLUMNS = [
    "module_name",
    "question_type",
    "question_text",
    "marks",
    "option_1",
    "option_2",
    "option_3",
    "option_4",
    "correct_option",
    "correct_answer_text",
]
_REQUIRED_TEMPLATE_COLUMNS = ("module_name", "question_type", "question_text")
_MAX_TEMPLATE_OPTIONS = 4


def _parse_import_rows(filename: str, content: bytes) -> tuple[list[tuple[int, dict]], str | None]:
    """Returns (rows, error). `rows` is a list of (spreadsheet_row_number, {header: value})
    - row numbers start at 2, matching what the user sees when they open the file, since
    row 1 is the header. `error` is set (rows is then []) when the file can't be read at
    all or is missing a required column - individual bad data rows are NOT an error here,
    they're reported per-row by the caller instead."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        import csv
        import io

        text = content.decode("utf-8-sig")
        raw_rows = list(csv.reader(io.StringIO(text)))
    elif ext in ("xlsx", "xlsm"):
        import io

        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.worksheets[0]
        raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        return [], "Unsupported file type - please upload a .csv or .xlsx file (use the downloadable template)."

    raw_rows = [r for r in raw_rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not raw_rows:
        return [], "The file is empty."

    headers = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]
    missing = [c for c in _REQUIRED_TEMPLATE_COLUMNS if c not in headers]
    if missing:
        return [], f"Missing required column(s): {', '.join(missing)}. Please use the downloaded template."

    rows = []
    for i, raw in enumerate(raw_rows[1:], start=2):
        row = {headers[j]: (raw[j] if j < len(raw) else None) for j in range(len(headers))}
        rows.append((i, row))
    return rows, None


def _cell_str(value) -> str:
    return str(value).strip() if value is not None else ""


def _row_to_question_request(row: dict) -> BankQuestionCreateRequest:
    module_name = _cell_str(row.get("module_name"))
    question_type = _cell_str(row.get("question_type")).upper()
    question_text = _cell_str(row.get("question_text"))

    marks_raw = row.get("marks")
    if marks_raw in (None, ""):
        marks = 1
    else:
        try:
            marks = int(marks_raw)
        except (TypeError, ValueError):
            raise ValueError(f"marks must be a whole number, got {marks_raw!r}") from None

    correct_option_raw = row.get("correct_option")
    correct_option_num = None
    if _cell_str(correct_option_raw) != "":
        try:
            correct_option_num = int(correct_option_raw)
        except (TypeError, ValueError):
            raise ValueError(f"correct_option must be a number, got {correct_option_raw!r}") from None

    options: list[OptionInput] = []
    correct_answer_text: str | None = None

    if question_type == QuestionType.MULTIPLE_CHOICE.value:
        filled = [
            (i, _cell_str(row.get(f"option_{i}")))
            for i in range(1, _MAX_TEMPLATE_OPTIONS + 1)
            if _cell_str(row.get(f"option_{i}"))
        ]
        if correct_option_num is not None and correct_option_num not in [i for i, _ in filled]:
            raise ValueError(f"correct_option {correct_option_num} does not match a filled option_{correct_option_num} cell")
        options = [OptionInput(option_text=text, is_correct=(i == correct_option_num)) for i, text in filled]
    else:
        correct_answer_text = _cell_str(row.get("correct_answer_text")) or None

    try:
        return BankQuestionCreateRequest(
            module_name=module_name,
            question_type=question_type,
            question_text=question_text,
            marks=marks,
            correct_answer_text=correct_answer_text,
            options=options,
        )
    except PydanticValidationError as e:
        raise ValueError("; ".join(err["msg"] for err in e.errors())) from None


async def import_questions(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, filename: str, content: bytes
) -> BankQuestionImportResponse | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    if bank is None or not _can_manage(current_user, bank.created_by):
        return None

    rows, file_error = _parse_import_rows(filename, content)
    if file_error:
        return BankQuestionImportResponse(created=0, errors=[BankQuestionImportError(row=1, message=file_error)])

    now = datetime.utcnow()
    created = 0
    errors: list[BankQuestionImportError] = []

    for row_number, row in rows:
        try:
            payload = _row_to_question_request(row)
        except ValueError as e:
            errors.append(BankQuestionImportError(row=row_number, message=str(e)))
            continue

        entity = BankQuestionEntity(
            bank_id=bank_id,
            module_name=payload.module_name,
            question_type=payload.question_type,
            question_text=payload.question_text,
            marks=payload.marks,
            correct_answer_text=payload.correct_answer_text,
            is_active=True,
            created_by=current_user.id,
            created_at=now,
            updated_at=now,
        )
        hrms_db.add(entity)
        await hrms_db.flush()
        for idx, opt in enumerate(payload.options):
            hrms_db.add(
                BankQuestionOptionEntity(
                    question_id=entity.id, option_text=opt.option_text, is_correct=opt.is_correct, display_order=idx
                )
            )
        created += 1

    await hrms_db.commit()
    return BankQuestionImportResponse(created=created, errors=errors)


def build_question_template_xlsx() -> bytes:
    import io

    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Questions"
    sheet.append(IMPORT_TEMPLATE_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    example_rows = [
        ["Python Basics", "MULTIPLE_CHOICE", "Which of the following is a mutable data type in Python?", 2,
         "Tuple", "List", "String", "Frozenset", 2, None],
        ["Python Basics", "TRUE_FALSE", "Python is a statically typed language.", 1,
         None, None, None, None, None, "False"],
        ["Python Basics", "FILL_IN_BLANK", "The ____ keyword is used to define a function in Python.", 1,
         None, None, None, None, None, "def"],
        ["Python Basics", "ENTER_ANSWER", "Name the Python package manager commonly used to install third-party libraries.", 1,
         None, None, None, None, None, "pip"],
    ]
    for row in example_rows:
        sheet.append(row)

    for col_cells in sheet.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 50)

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Column", "Notes"])
    for cell in instructions[1]:
        cell.font = Font(bold=True)
    for note in [
        ("module_name", "Required. Free-text grouping label used to filter questions when building a Task."),
        ("question_type", "Required. One of: MULTIPLE_CHOICE, TRUE_FALSE, FILL_IN_BLANK, ENTER_ANSWER."),
        ("question_text", "Required."),
        ("marks", "Optional whole number, defaults to 1 if left blank."),
        ("option_1 .. option_4", "MULTIPLE_CHOICE only. Fill in 2-4 options; leave any unused ones blank."),
        ("correct_option", "MULTIPLE_CHOICE only. The number (1-4) of the correct option's column."),
        (
            "correct_answer_text",
            "TRUE_FALSE / FILL_IN_BLANK / ENTER_ANSWER only. e.g. True, False, or the expected answer "
            "(matched case-insensitively, ignoring leading/trailing spaces).",
        ),
        ("", ""),
        ("Tip", "Delete the 4 example rows before uploading, or leave them out - they'll be imported as real questions otherwise."),
    ]:
        instructions.append(list(note))
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 95

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def list_questions(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, module_name: str | None, page_number: int, page_size: int
) -> PageResult | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    if bank is None:
        return None
    stmt = select(BankQuestionEntity).where(BankQuestionEntity.bank_id == bank_id, BankQuestionEntity.is_active.is_(True))
    if module_name:
        stmt = stmt.where(BankQuestionEntity.module_name == module_name)
    stmt = stmt.order_by(BankQuestionEntity.id.desc())
    page_result = await paginate(hrms_db, stmt, page_number, page_size)
    page_result.items = [await _to_bank_question_response(hrms_db, q) for q in page_result.items]
    return page_result


async def list_module_names(hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int) -> list[str] | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    if bank is None:
        return None
    result = await hrms_db.execute(
        select(BankQuestionEntity.module_name).where(BankQuestionEntity.bank_id == bank_id).distinct()
    )
    return sorted({row[0] for row in result.all()})


async def get_question(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, question_id: int
) -> BankQuestionResponse | None:
    entity = await hrms_db.get(BankQuestionEntity, question_id)
    if entity is None or entity.bank_id != bank_id:
        return None
    return await _to_bank_question_response(hrms_db, entity)


async def update_question(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, question_id: int, data: BankQuestionUpdateRequest
) -> BankQuestionResponse | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    entity = await hrms_db.get(BankQuestionEntity, question_id)
    if bank is None or entity is None or entity.bank_id != bank_id or not _can_manage(current_user, bank.created_by):
        return None

    entity.module_name = data.module_name
    entity.question_type = data.question_type
    entity.question_text = data.question_text
    entity.marks = data.marks
    entity.correct_answer_text = data.correct_answer_text
    entity.updated_at = datetime.utcnow()

    existing_options = (
        await hrms_db.execute(select(BankQuestionOptionEntity).where(BankQuestionOptionEntity.question_id == entity.id))
    ).scalars().all()
    for opt in existing_options:
        await hrms_db.delete(opt)
    await hrms_db.flush()
    for idx, opt in enumerate(data.options):
        hrms_db.add(
            BankQuestionOptionEntity(question_id=entity.id, option_text=opt.option_text, is_correct=opt.is_correct, display_order=idx)
        )

    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_question_response(hrms_db, entity)


async def set_question_active(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, bank_id: int, question_id: int, is_active: bool
) -> BankQuestionResponse | None:
    bank = await hrms_db.get(QuestionBankEntity, bank_id)
    entity = await hrms_db.get(BankQuestionEntity, question_id)
    if bank is None or entity is None or entity.bank_id != bank_id or not _can_manage(current_user, bank.created_by):
        return None
    entity.is_active = is_active
    entity.updated_at = datetime.utcnow()
    await hrms_db.commit()
    await hrms_db.refresh(entity)
    return await _to_bank_question_response(hrms_db, entity)


# ---------------------------------------------------------------------------
# Tasks
# ---------------------------------------------------------------------------


async def _snapshot_bank_questions(
    hrms_db: AsyncSession, task_id: int, assignee_id: int | None, questions: list[BankQuestionEntity], now: datetime
) -> None:
    """Freezes `questions` (in the given order) onto TaskQuestionEntity/
    TaskQuestionOptionEntity rows. assignee_id=None makes a shared MANUAL-mode snapshot
    (task-creation time, one set for every assignee); a specific assignee_id makes a
    RANDOM-mode per-attempt snapshot instead (see _snapshot_random_questions). Does not
    commit - the caller does, as part of a larger unit of work."""
    mcq_ids = [q.id for q in questions if q.question_type == QuestionType.MULTIPLE_CHOICE.value]
    options_by_question: dict[int, list[BankQuestionOptionEntity]] = {}
    if mcq_ids:
        opts_result = await hrms_db.execute(
            select(BankQuestionOptionEntity)
            .where(BankQuestionOptionEntity.question_id.in_(mcq_ids))
            .order_by(BankQuestionOptionEntity.display_order, BankQuestionOptionEntity.id)
        )
        for o in opts_result.scalars().all():
            options_by_question.setdefault(o.question_id, []).append(o)

    for idx, q in enumerate(questions):
        task_question = TaskQuestionEntity(
            task_id=task_id,
            assignee_id=assignee_id,
            source_question_id=q.id,
            module_name=q.module_name,
            question_type=q.question_type,
            question_text=q.question_text,
            correct_answer_text=q.correct_answer_text,
            marks=q.marks,
            display_order=idx,
            created_at=now,
        )
        hrms_db.add(task_question)
        await hrms_db.flush()
        for opt in options_by_question.get(q.id, []):
            hrms_db.add(
                TaskQuestionOptionEntity(
                    task_question_id=task_question.id, option_text=opt.option_text, is_correct=opt.is_correct, display_order=opt.display_order
                )
            )


async def _snapshot_random_questions(hrms_db: AsyncSession, task: TaskEntity, assignee: TaskAssigneeEntity) -> None:
    """Called once, the first time a given attempt is started (see start_or_resume) -
    independently draws task.random_question_count active questions from
    task.source_bank_id(+source_module_name) for THIS assignee/attempt only, so a retry
    also gets a fresh draw."""
    stmt = select(BankQuestionEntity).where(
        BankQuestionEntity.bank_id == task.source_bank_id, BankQuestionEntity.is_active.is_(True)
    )
    if task.source_module_name:
        stmt = stmt.where(BankQuestionEntity.module_name == task.source_module_name)
    pool = (await hrms_db.execute(stmt)).scalars().all()
    if len(pool) < task.random_question_count:
        raise ValueError(
            f"The question bank no longer has enough active questions for this task's random draw "
            f"({len(pool)} available, {task.random_question_count} needed)"
        )
    chosen = random.sample(list(pool), task.random_question_count)
    await _snapshot_bank_questions(hrms_db, task.id, assignee.id, chosen, datetime.utcnow())
    await hrms_db.commit()


async def create_task(hrms_db: AsyncSession, current_user: CurrentHrmsUser, data: TaskCreateRequest) -> TaskManageDetailResponse:
    questions_by_id: dict[int, BankQuestionEntity] = {}
    if data.question_mode == QuestionMode.RANDOM.value:
        bank = await hrms_db.get(QuestionBankEntity, data.source_bank_id)
        if bank is None or not bank.is_active:
            raise ValueError("Question bank not found or inactive")
        count_stmt = (
            select(func.count())
            .select_from(BankQuestionEntity)
            .where(BankQuestionEntity.bank_id == bank.id, BankQuestionEntity.is_active.is_(True))
        )
        if data.source_module_name:
            count_stmt = count_stmt.where(BankQuestionEntity.module_name == data.source_module_name)
        available = (await hrms_db.execute(count_stmt)).scalar_one()
        if available < data.random_question_count:
            raise ValueError(
                f"This bank only has {available} active question(s) matching the filter - "
                f"need at least {data.random_question_count} for a random draw of that size"
            )
    else:
        q_result = await hrms_db.execute(
            select(BankQuestionEntity).where(BankQuestionEntity.id.in_(data.question_ids), BankQuestionEntity.is_active.is_(True))
        )
        questions_by_id = {q.id: q for q in q_result.scalars().all()}
        missing = set(data.question_ids) - set(questions_by_id)
        if missing:
            raise ValueError(f"Question(s) not found or inactive: {sorted(missing)}")

    trainee_ids = sorted(set(data.trainee_ids))
    trainee_users_by_id: dict[int, HrmsUserEntity] = {}
    if trainee_ids:
        # Only HR/Admin may assign trainees - a Trainer (TEAM_MEMBER) authoring a task
        # must leave trainee_ids empty; HR/Admin assigns later via assign_trainees.
        if current_user.role not in (Role.ADMIN, Role.HR):
            raise ValueError("Only HR/Admin can assign trainees to a task")
        t_result = await hrms_db.execute(select(HrmsUserEntity).where(HrmsUserEntity.id.in_(trainee_ids)))
        trainee_users_by_id = {u.id: u for u in t_result.scalars().all()}
        for trainee_id in trainee_ids:
            trainee = trainee_users_by_id.get(trainee_id)
            if trainee is None or trainee.role != Role.TEAM_MEMBER:
                raise ValueError(f"Trainee {trainee_id} must be a Team Member")

    if data.training_id is not None:
        training = await hrms_db.get(TrainingProgramEntity, data.training_id)
        if training is None:
            raise ValueError("Training not found")
        if training.status != TrainingStatus.COMPLETED.value:
            raise ValueError("The training must be completed before HR can give its assessment")
        if training.assessment_given_by != AssessmentGivenBy.HR.value:
            raise ValueError("This training's assessment is given by the Trainer, not HR")
        if trainee_ids:
            valid_trainee_ids = await _training_trainee_ids(hrms_db, data.training_id)
            invalid = set(trainee_ids) - valid_trainee_ids
            if invalid:
                raise ValueError(f"Trainee(s) {sorted(invalid)} are not part of this training")

    now = datetime.utcnow()
    task = TaskEntity(
        title=data.title,
        description=data.description,
        training_id=data.training_id,
        time_limit_minutes=data.time_limit_minutes,
        pass_percentage=data.pass_percentage,
        max_attempts=data.max_attempts,
        question_mode=data.question_mode,
        source_bank_id=data.source_bank_id,
        source_module_name=data.source_module_name,
        random_question_count=data.random_question_count,
        difficulty_level=data.difficulty_level,
        skill_name=data.skill_name,
        skill_category=data.skill_category,
        status=TaskStatus.OPEN.value,
        created_by=current_user.id,
        created_at=now,
        updated_at=now,
    )
    hrms_db.add(task)
    await hrms_db.flush()

    if data.question_mode == QuestionMode.MANUAL.value:
        # Shared snapshot (assignee_id=None) taken once now - every assignee reads the
        # same rows. A RANDOM-mode task instead snapshots independently per assignee,
        # the first time each attempt is started (see _snapshot_random_questions).
        await _snapshot_bank_questions(hrms_db, task.id, None, [questions_by_id[qid] for qid in data.question_ids], now)

    for trainee_id in trainee_ids:
        hrms_db.add(
            TaskAssigneeEntity(
                task_id=task.id,
                trainee_id=trainee_id,
                attempt_number=1,
                max_attempts=task.max_attempts,
                status=AssigneeStatus.NOT_STARTED.value,
                created_at=now,
            )
        )

    await hrms_db.commit()

    for trainee_id in trainee_ids:
        trainee = trainee_users_by_id[trainee_id]
        await email_service.send_task_assigned_email(trainee.email, trainee.name, task.title, task.time_limit_minutes, task.id)

    return await get_task_for_management(hrms_db, current_user, task.id)


async def assign_trainees(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int, trainee_ids: list[int]
) -> TaskManageDetailResponse | None:
    """Route-level require_role(ADMIN, HR) already restricts callers - this is
    deliberately NOT ownership-gated, so HR/Admin can assign to any task regardless of
    who authored it (including one a Trainer built with no assignees)."""
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    if task.status != TaskStatus.OPEN.value:
        raise ValueError("Cannot assign trainees to a closed task")

    existing_result = await hrms_db.execute(select(TaskAssigneeEntity.trainee_id).where(TaskAssigneeEntity.task_id == task_id))
    already_assigned = {row[0] for row in existing_result.all()}
    new_ids = sorted({tid for tid in trainee_ids if tid not in already_assigned})
    if new_ids:
        t_result = await hrms_db.execute(select(HrmsUserEntity).where(HrmsUserEntity.id.in_(new_ids)))
        users_by_id = {u.id: u for u in t_result.scalars().all()}
        for trainee_id in new_ids:
            trainee = users_by_id.get(trainee_id)
            if trainee is None or trainee.role != Role.TEAM_MEMBER:
                raise ValueError(f"Trainee {trainee_id} must be a Team Member")

        if task.training_id is not None:
            valid_trainee_ids = await _training_trainee_ids(hrms_db, task.training_id)
            invalid = set(new_ids) - valid_trainee_ids
            if invalid:
                raise ValueError(f"Trainee(s) {sorted(invalid)} are not part of this training")

        now = datetime.utcnow()
        for trainee_id in new_ids:
            hrms_db.add(
                TaskAssigneeEntity(
                    task_id=task_id,
                    trainee_id=trainee_id,
                    attempt_number=1,
                    max_attempts=task.max_attempts,
                    status=AssigneeStatus.NOT_STARTED.value,
                    created_at=now,
                )
            )
        await hrms_db.commit()

        for trainee_id in new_ids:
            trainee = users_by_id[trainee_id]
            await email_service.send_task_assigned_email(trainee.email, trainee.name, task.title, task.time_limit_minutes, task.id)

    return await get_task_for_management(hrms_db, current_user, task_id)


async def assign_trainees_by_location(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int, work_location: str
) -> TaskManageDetailResponse | None:
    """Bulk-assign equivalent of assign_trainees: resolves every Team Member at the
    given work location (see hrms_settings.work_locations) and assigns them all in one
    call, rather than requiring the caller to pick trainees one by one."""
    trainee_ids = await user_service.list_team_member_ids_by_location(hrms_db, work_location)
    if not trainee_ids:
        raise ValueError(f"No employees found at work location '{work_location}'")
    return await assign_trainees(hrms_db, current_user, task_id, trainee_ids)


async def list_tasks(hrms_db: AsyncSession, current_user: CurrentHrmsUser, page_number: int, page_size: int) -> PageResult:
    if current_user.role in (Role.ADMIN, Role.HR):
        stmt = select(TaskEntity)
    else:
        assignee_subq = select(TaskAssigneeEntity.task_id).where(TaskAssigneeEntity.trainee_id == current_user.id)
        stmt = select(TaskEntity).where(or_(TaskEntity.created_by == current_user.id, TaskEntity.id.in_(assignee_subq)))
    stmt = stmt.order_by(TaskEntity.id.desc())
    page_result = await paginate(hrms_db, stmt, page_number, page_size)

    tasks = page_result.items
    task_ids = [t.id for t in tasks]
    tasks_by_id = {t.id: t for t in tasks}
    names = await _user_names(hrms_db, {t.created_by for t in tasks})

    question_counts: dict[int, int] = {}
    total_marks_by_task: dict[int, int] = {}
    if task_ids:
        # Only the shared (assignee_id IS NULL) MANUAL-mode snapshot is counted here - a
        # RANDOM-mode task has no such rows and gets its count/marks overridden below,
        # since summing per-assignee draws would otherwise multiply by assignee count.
        q_result = await hrms_db.execute(
            select(TaskQuestionEntity.task_id, func.count(), func.sum(TaskQuestionEntity.marks))
            .where(TaskQuestionEntity.task_id.in_(task_ids), TaskQuestionEntity.assignee_id.is_(None))
            .group_by(TaskQuestionEntity.task_id)
        )
        for task_id, count, total in q_result.all():
            question_counts[task_id] = count
            total_marks_by_task[task_id] = total or 0
    for t in tasks:
        if t.question_mode == QuestionMode.RANDOM.value:
            question_counts[t.id] = t.random_question_count or 0
            total_marks_by_task[t.id] = 0

    assignee_counts: dict[int, int] = {}
    submitted_counts: dict[int, int] = {}
    my_assignees: dict[int, TaskAssigneeEntity] = {}
    latest_by_task = await _latest_assignees_for_tasks(hrms_db, task_ids)
    for task_id, assignees in latest_by_task.items():
        assignee_counts[task_id] = len(assignees)
        for a in assignees:
            if a.status in _FINALIZED_STATUSES:
                submitted_counts[task_id] = submitted_counts.get(task_id, 0) + 1
            if a.trainee_id == current_user.id:
                my_assignees[task_id] = a

    for task_id, assignee in my_assignees.items():
        if assignee.status == AssigneeStatus.IN_PROGRESS.value:
            await _finalize_if_expired(hrms_db, assignee, tasks_by_id[task_id])
            if assignee.status in _FINALIZED_STATUSES:
                submitted_counts[task_id] = submitted_counts.get(task_id, 0) + 1

    page_result.items = [
        {
            "id": t.id,
            "title": t.title,
            "status": t.status,
            "training_id": t.training_id,
            "time_limit_minutes": t.time_limit_minutes,
            "pass_percentage": t.pass_percentage,
            "max_attempts": t.max_attempts,
            "question_mode": t.question_mode,
            "difficulty_level": t.difficulty_level,
            "skill_name": t.skill_name,
            "skill_category": t.skill_category,
            "created_by": t.created_by,
            "created_by_name": names.get(t.created_by, "Unknown"),
            "question_count": question_counts.get(t.id, 0),
            "total_marks": total_marks_by_task.get(t.id, 0),
            "assignee_count": assignee_counts.get(t.id, 0),
            "submitted_count": submitted_counts.get(t.id, 0),
            "created_at": t.created_at,
            "my_assignee": (
                {
                    "id": my_assignees[t.id].id,
                    "status": my_assignees[t.id].status,
                    "started_at": my_assignees[t.id].started_at,
                    "deadline_at": my_assignees[t.id].deadline_at,
                    "submitted_at": my_assignees[t.id].submitted_at,
                    "percentage": my_assignees[t.id].percentage,
                    "passed": my_assignees[t.id].passed,
                    "attempt_number": my_assignees[t.id].attempt_number,
                    "max_attempts": my_assignees[t.id].max_attempts,
                    "can_retry": _assignee_attempt_flags(my_assignees[t.id])[0],
                    "locked_out": _assignee_attempt_flags(my_assignees[t.id])[1],
                }
                if t.id in my_assignees
                else None
            ),
        }
        for t in tasks
    ]
    return page_result


async def _task_questions_with_options(
    hrms_db: AsyncSession, task_id: int, assignee_id: int | None = None
) -> tuple[list[TaskQuestionEntity], dict[int, list[TaskQuestionOptionEntity]]]:
    """assignee_id=None (the default) returns only the shared MANUAL-mode snapshot
    (rows with assignee_id IS NULL) - used for the management view, where a RANDOM-mode
    task correctly comes back empty since it has no single shared question set. Passing
    a specific assignee_id additionally includes that attempt's own RANDOM-mode draw."""
    condition = (
        TaskQuestionEntity.assignee_id.is_(None)
        if assignee_id is None
        else or_(TaskQuestionEntity.assignee_id.is_(None), TaskQuestionEntity.assignee_id == assignee_id)
    )
    tq_result = await hrms_db.execute(
        select(TaskQuestionEntity)
        .where(TaskQuestionEntity.task_id == task_id, condition)
        .order_by(TaskQuestionEntity.display_order, TaskQuestionEntity.id)
    )
    task_questions = list(tq_result.scalars().all())
    tq_ids = [q.id for q in task_questions]
    options_by_question: dict[int, list[TaskQuestionOptionEntity]] = {}
    if tq_ids:
        opts_result = await hrms_db.execute(
            select(TaskQuestionOptionEntity)
            .where(TaskQuestionOptionEntity.task_question_id.in_(tq_ids))
            .order_by(TaskQuestionOptionEntity.display_order, TaskQuestionOptionEntity.id)
        )
        for o in opts_result.scalars().all():
            options_by_question.setdefault(o.task_question_id, []).append(o)
    return task_questions, options_by_question


async def get_task_for_management(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int
) -> TaskManageDetailResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None or not _can_manage(current_user, task.created_by):
        return None

    task_questions, options_by_question = await _task_questions_with_options(hrms_db, task_id)
    total_marks = sum(q.marks for q in task_questions)

    assignees = await _latest_assignees_for_task(hrms_db, task_id)
    for a in assignees:
        if a.status == AssigneeStatus.IN_PROGRESS.value:
            await _finalize_if_expired(hrms_db, a, task)
    names = await _user_names(hrms_db, {a.trainee_id for a in assignees} | {task.created_by})

    source_bank_name = None
    if task.source_bank_id:
        source_bank = await hrms_db.get(QuestionBankEntity, task.source_bank_id)
        source_bank_name = source_bank.name if source_bank else None

    return TaskManageDetailResponse(
        id=task.id,
        title=task.title,
        description=task.description,
        status=task.status,
        training_id=task.training_id,
        time_limit_minutes=task.time_limit_minutes,
        pass_percentage=task.pass_percentage,
        max_attempts=task.max_attempts,
        question_mode=task.question_mode,
        source_bank_id=task.source_bank_id,
        source_bank_name=source_bank_name,
        source_module_name=task.source_module_name,
        random_question_count=task.random_question_count,
        difficulty_level=task.difficulty_level,
        skill_name=task.skill_name,
        skill_category=task.skill_category,
        created_by=task.created_by,
        created_by_name=names.get(task.created_by, "Unknown"),
        total_marks=total_marks,
        questions=[
            TaskQuestionManageItem(
                id=q.id,
                module_name=q.module_name,
                question_type=q.question_type,
                question_text=q.question_text,
                marks=q.marks,
                correct_answer_text=q.correct_answer_text,
                display_order=q.display_order,
                options=[
                    TaskOptionManageItem(id=o.id, option_text=o.option_text, is_correct=o.is_correct, display_order=o.display_order)
                    for o in options_by_question.get(q.id, [])
                ],
            )
            for q in task_questions
        ],
        assignees=[
            TaskAssigneeSummary(
                id=a.id,
                trainee_id=a.trainee_id,
                trainee_name=names.get(a.trainee_id, "Unknown"),
                status=a.status,
                started_at=a.started_at,
                deadline_at=a.deadline_at,
                submitted_at=a.submitted_at,
                submit_reason=a.submit_reason,
                marks_obtained=a.marks_obtained,
                total_marks=a.total_marks,
                percentage=a.percentage,
                passed=a.passed,
                attempt_number=a.attempt_number,
                max_attempts=a.max_attempts,
                can_retry=_assignee_attempt_flags(a)[0],
                locked_out=_assignee_attempt_flags(a)[1],
            )
            for a in assignees
        ],
        created_at=task.created_at,
        updated_at=task.updated_at,
        closed_at=task.closed_at,
    )


async def list_tasks_for_training(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, training_id: int
) -> list[TaskManageDetailResponse]:
    """The HR-given Task Assessments (one per trainee, typically) linked to a specific
    Training. ADMIN/HR only - reuses get_task_for_management per task, so results are
    identical in shape/ownership-gating to fetching each task individually."""
    result = await hrms_db.execute(
        select(TaskEntity.id).where(TaskEntity.training_id == training_id).order_by(TaskEntity.id)
    )
    task_ids = [row[0] for row in result.all()]
    tasks = [await get_task_for_management(hrms_db, current_user, task_id) for task_id in task_ids]
    return [t for t in tasks if t is not None]


async def get_task_for_trainee(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> TaskTakeResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None
    await _finalize_if_expired(hrms_db, assignee, task)

    task_questions, options_by_question = await _task_questions_with_options(hrms_db, task_id, assignee.id)
    tq_ids = [q.id for q in task_questions]

    my_answers: dict[int, TaskAnswerEntity] = {}
    if tq_ids:
        ans_result = await hrms_db.execute(
            select(TaskAnswerEntity).where(TaskAnswerEntity.assignee_id == assignee.id, TaskAnswerEntity.task_question_id.in_(tq_ids))
        )
        for ans in ans_result.scalars().all():
            my_answers[ans.task_question_id] = ans

    return TaskTakeResponse(
        id=task.id,
        title=task.title,
        description=task.description,
        status=task.status,
        time_limit_minutes=task.time_limit_minutes,
        started_at=assignee.started_at,
        deadline_at=assignee.deadline_at,
        attempt_number=assignee.attempt_number,
        max_attempts=assignee.max_attempts,
        questions=[
            TaskQuestionTakeItem(
                id=q.id,
                module_name=q.module_name,
                question_type=q.question_type,
                question_text=q.question_text,
                marks=q.marks,
                display_order=q.display_order,
                options=(
                    [
                        TaskOptionTakeItem(id=o.id, option_text=o.option_text, display_order=o.display_order)
                        for o in options_by_question.get(q.id, [])
                    ]
                    if q.question_type == QuestionType.MULTIPLE_CHOICE.value
                    else None
                ),
                my_answer=my_answers[q.id].answer_text if q.id in my_answers else None,
                my_selected_option_id=my_answers[q.id].selected_option_id if q.id in my_answers else None,
            )
            for q in task_questions
        ],
    )


async def close_task(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> TaskManageDetailResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None or not _can_manage(current_user, task.created_by):
        return None
    if task.status != TaskStatus.CLOSED.value:
        now = datetime.utcnow()
        assignees = await _latest_assignees_for_task(hrms_db, task_id)
        for assignee in assignees:
            if assignee.status == AssigneeStatus.IN_PROGRESS.value:
                await _grade_and_finalize(hrms_db, assignee, task, AssigneeStatus.AUTO_SUBMITTED.value, now)

        task.status = TaskStatus.CLOSED.value
        task.closed_at = now
        task.updated_at = now
        await hrms_db.commit()
    return await get_task_for_management(hrms_db, current_user, task_id)


async def start_or_resume(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> TaskTakeResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None
    if assignee.status == AssigneeStatus.NOT_STARTED.value:
        if task.status != TaskStatus.OPEN.value:
            raise ValueError("This task is closed and can no longer be started")
        if task.question_mode == QuestionMode.RANDOM.value:
            await _snapshot_random_questions(hrms_db, task, assignee)
        now = datetime.utcnow()
        assignee.status = AssigneeStatus.IN_PROGRESS.value
        assignee.started_at = now
        assignee.deadline_at = now + timedelta(minutes=task.time_limit_minutes)
        await hrms_db.commit()
    return await get_task_for_trainee(hrms_db, current_user, task_id)


async def save_answer(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int, task_question_id: int, data: AnswerSaveRequest
) -> AnswerSaveResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None

    await _finalize_if_expired(hrms_db, assignee, task)
    if assignee.status != AssigneeStatus.IN_PROGRESS.value:
        raise ValueError("This task is no longer in progress (time may already be up)")

    task_question = await hrms_db.get(TaskQuestionEntity, task_question_id)
    if (
        task_question is None
        or task_question.task_id != task_id
        or (task_question.assignee_id is not None and task_question.assignee_id != assignee.id)
    ):
        raise ValueError("Question not found on this task")

    options_by_id: dict[int, TaskQuestionOptionEntity] = {}
    if task_question.question_type == QuestionType.MULTIPLE_CHOICE.value:
        opts_result = await hrms_db.execute(select(TaskQuestionOptionEntity).where(TaskQuestionOptionEntity.task_question_id == task_question_id))
        options_by_id = {o.id: o for o in opts_result.scalars().all()}
        if data.selected_option_id is not None and data.selected_option_id not in options_by_id:
            raise ValueError("selected_option_id does not belong to this question")

    is_correct, marks_awarded = _grade_single_answer(task_question, options_by_id, data.selected_option_id, data.answer_text)

    existing_result = await hrms_db.execute(
        select(TaskAnswerEntity).where(TaskAnswerEntity.assignee_id == assignee.id, TaskAnswerEntity.task_question_id == task_question_id)
    )
    answer = existing_result.scalar_one_or_none()
    if answer is None:
        answer = TaskAnswerEntity(assignee_id=assignee.id, task_question_id=task_question_id)
        hrms_db.add(answer)
    answer.selected_option_id = data.selected_option_id
    answer.answer_text = data.answer_text
    answer.is_correct = is_correct
    answer.marks_awarded = marks_awarded
    answer.answered_at = datetime.utcnow()

    await hrms_db.commit()
    return AnswerSaveResponse(task_question_id=task_question_id, saved=True)


async def submit_task(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int, reason: str | None = None
) -> TaskResultResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None

    await _finalize_if_expired(hrms_db, assignee, task)
    if assignee.status == AssigneeStatus.NOT_STARTED.value:
        raise ValueError("Task has not been started yet")
    if assignee.status == AssigneeStatus.IN_PROGRESS.value:
        assignee.submit_reason = reason
    await _grade_and_finalize(hrms_db, assignee, task, AssigneeStatus.SUBMITTED.value, datetime.utcnow())

    return _to_result_response(task.id, assignee)


async def get_my_result(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> TaskResultResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None
    await _finalize_if_expired(hrms_db, assignee, task)

    return _to_result_response(task.id, assignee)


async def retry_task(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> RetryResponse | None:
    """Trainee-initiated: only available after a FAILED, finalized attempt with attempts
    still remaining (see _assignee_attempt_flags) - creates the next attempt row at the
    same max_attempts cap. Once attempt_number reaches max_attempts, the trainee is
    locked_out instead and only HR's reassign_task can grant more."""
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, current_user.id)
    if assignee is None:
        return None
    await _finalize_if_expired(hrms_db, assignee, task)

    can_retry, _ = _assignee_attempt_flags(assignee)
    if not can_retry:
        raise ValueError("No retry is available for this task")
    if task.status != TaskStatus.OPEN.value:
        raise ValueError("This task is closed and can no longer be retried")

    now = datetime.utcnow()
    new_assignee = TaskAssigneeEntity(
        task_id=task_id,
        trainee_id=current_user.id,
        attempt_number=assignee.attempt_number + 1,
        max_attempts=assignee.max_attempts,
        status=AssigneeStatus.NOT_STARTED.value,
        created_at=now,
    )
    hrms_db.add(new_assignee)
    await hrms_db.commit()
    await hrms_db.refresh(new_assignee)
    return RetryResponse(attempt_number=new_assignee.attempt_number, max_attempts=new_assignee.max_attempts, status=new_assignee.status)


async def reassign_task(
    hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int, trainee_id: int
) -> TaskManageDetailResponse | None:
    """HR/Admin-only (route-gated, not ownership-gated - same precedent as
    assign_trainees). Only valid once a trainee is locked_out (every attempt of their
    current cycle failed) - grants a fresh cycle by bumping max_attempts by another
    +task.max_attempts and starting the next attempt_number at NOT_STARTED, without
    touching any other trainee assigned to this same task."""
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None:
        return None
    assignee = await _get_latest_assignee(hrms_db, task_id, trainee_id)
    if assignee is None:
        raise ValueError("This trainee is not assigned to this task")
    await _finalize_if_expired(hrms_db, assignee, task)

    _, locked_out = _assignee_attempt_flags(assignee)
    if not locked_out:
        raise ValueError("This trainee has not exhausted all attempts yet - reassign is only for a locked-out trainee")
    if task.status != TaskStatus.OPEN.value:
        raise ValueError("This task is closed and can no longer be reassigned")

    now = datetime.utcnow()
    new_assignee = TaskAssigneeEntity(
        task_id=task_id,
        trainee_id=trainee_id,
        attempt_number=assignee.attempt_number + 1,
        max_attempts=assignee.max_attempts + task.max_attempts,
        status=AssigneeStatus.NOT_STARTED.value,
        created_at=now,
    )
    hrms_db.add(new_assignee)
    await hrms_db.commit()
    return await get_task_for_management(hrms_db, current_user, task_id)


async def get_report(hrms_db: AsyncSession, current_user: CurrentHrmsUser, task_id: int) -> TaskReportResponse | None:
    task = await hrms_db.get(TaskEntity, task_id)
    if task is None or not _can_manage(current_user, task.created_by):
        return None

    assignees = await _latest_assignees_for_task(hrms_db, task_id)
    for a in assignees:
        if a.status == AssigneeStatus.IN_PROGRESS.value:
            await _finalize_if_expired(hrms_db, a, task)

    names = await _user_names(hrms_db, {a.trainee_id for a in assignees})

    not_started = sum(1 for a in assignees if a.status == AssigneeStatus.NOT_STARTED.value)
    in_progress = sum(1 for a in assignees if a.status == AssigneeStatus.IN_PROGRESS.value)
    submitted = [a for a in assignees if a.status in _FINALIZED_STATUSES]
    pass_count = sum(1 for a in submitted if a.passed)
    average_percentage = round(sum(a.percentage or 0 for a in submitted) / len(submitted), 2) if submitted else None

    return TaskReportResponse(
        task_id=task.id,
        title=task.title,
        pass_percentage=task.pass_percentage,
        assignee_count=len(assignees),
        not_started_count=not_started,
        in_progress_count=in_progress,
        submitted_count=len(submitted),
        pass_count=pass_count,
        average_percentage=average_percentage,
        rows=[
            TaskReportRow(
                assignee_id=a.id,
                trainee_id=a.trainee_id,
                trainee_name=names.get(a.trainee_id, "Unknown"),
                status=a.status,
                started_at=a.started_at,
                deadline_at=a.deadline_at,
                submitted_at=a.submitted_at,
                submit_reason=a.submit_reason,
                marks_obtained=a.marks_obtained,
                total_marks=a.total_marks,
                percentage=a.percentage,
                passed=a.passed,
                attempt_number=a.attempt_number,
                max_attempts=a.max_attempts,
                locked_out=_assignee_attempt_flags(a)[1],
            )
            for a in assignees
        ],
    )
